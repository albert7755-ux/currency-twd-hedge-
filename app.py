import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import requests
from datetime import datetime, timedelta

# ── Google Drive 設定 ──────────────────────────────────────
FUND_FOLDER_ID = "1i1-zUzLNnuwo2NVWijubvBICLbladZQO"

FUND_DB = {
    "F0HKG05X22_FO": "安聯台灣科技 (ACDD04)",
    "F00001EBH4_FO": "元大全球優質龍頭平衡 (ACYT168)",
    "F00001DRQQ_FO": "PIMCO收益增長",
    "F0GBR04SG1_FO": "駿利亨德森平衡基金",
    "F00000ZXFV_FO": "施羅德環球收息債券",
    "F00000PR1I_FO": "富達全球優質債券基金",
    "F000011JGT_FO": "群益潛力收益多重",
    "F0GBR04MRL_FO": "聯博美國收益EA穩定月配",
    "FOGBR05KHT_FO": "PIMCO多元收益",
    "F0GBR04AMK_FO": "貝萊德環球資產配置基金",
    "F00000MLER_FO": "聯博新興市場多元收益",
    "F00000T0K2_FO": "聯博美國成長基金EP",
    "F00000V557_FO": "聯博全球多元",
    "F00001EQPP_FO": "富邦台美雙星多重",
}

@st.cache_resource
def get_gspread_client():
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds_dict = json.loads(st.secrets["GOOGLE_CREDENTIALS"])
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly"
        ]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        return gspread.authorize(creds)
    except Exception as e:
        return None

def get_drive_headers():
    from google.oauth2.service_account import Credentials
    from google.auth.transport.requests import Request
    creds_dict = json.loads(st.secrets["GOOGLE_CREDENTIALS"])
    scopes = ["https://www.googleapis.com/auth/drive.readonly"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    creds.refresh(Request())
    return {"Authorization": f"Bearer {creds.token}"}

@st.cache_data(ttl=3600)
def list_sheets_in_folder(folder_id):
    try:
        headers = get_drive_headers()
        params = {
            "q": f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.spreadsheet' and trashed=false",
            "fields": "files(id, name)",
            "pageSize": 200,
        }
        resp = requests.get("https://www.googleapis.com/drive/v3/files", headers=headers, params=params)
        return {f["name"]: f["id"] for f in resp.json().get("files", [])}
    except:
        return {}

@st.cache_data(ttl=3600)
def read_nav_series(sheet_id, label):
    try:
        client = get_gspread_client()
        sh = client.open_by_key(sheet_id)
        ws = sh.get_worksheet(0)
        data = ws.get_all_records()
        df = pd.DataFrame(data)
        date_col = df.columns[0]
        val_col = df.columns[1]
        try:
            df["date"] = pd.to_datetime(df[date_col], unit="s", errors="coerce")
            if df["date"].isna().mean() > 0.5:
                df["date"] = pd.to_datetime(df[date_col], errors="coerce")
        except:
            df["date"] = pd.to_datetime(df[date_col], errors="coerce")
        df = df.dropna(subset=["date"])
        df = df.sort_values("date").set_index("date")
        return df[val_col].astype(float).rename(label)
    except Exception as e:
        return None

def calc_period_return(nav_series, period_years):
    """從淨值序列計算特定期間的累積總報酬%（不年化）"""
    end = nav_series.index[-1]
    days = int(period_years * 365)
    start = end - timedelta(days=days)
    subset = nav_series[nav_series.index >= start]
    if len(subset) < 5:
        return None
    ret = (subset.iloc[-1] / subset.iloc[0] - 1) * 100
    return round(ret, 2)

def fetch_fund_returns_from_drive(selected_tickers):
    """從 Google Drive 抓基金淨值並計算各期間累積報酬率"""
    fund_sheets = list_sheets_in_folder(FUND_FOLDER_ID)
    rows = []
    debug_info = []
    progress = st.progress(0, text="讀取基金資料中...")
    for i, ticker in enumerate(selected_tickers):
        fund_name = FUND_DB.get(ticker, ticker)
        sheet_id = fund_sheets.get(ticker)
        if not sheet_id:
            st.warning(f"⚠️ 找不到 {fund_name} 的試算表（ticker={ticker}），跳過")
            continue
        nav = read_nav_series(sheet_id, fund_name)
        if nav is None or len(nav) < 10:
            st.warning(f"⚠️ {fund_name} 資料不足，跳過")
            continue

        row = {"基金名稱": fund_name}
        fund_debug = {
            "基金": fund_name,
            "資料起始": nav.index[0].strftime("%Y-%m-%d"),
            "資料結束": nav.index[-1].strftime("%Y-%m-%d"),
            "起始淨值": round(float(nav.iloc[0]), 4),
            "結束淨值": round(float(nav.iloc[-1]), 4),
            "資料筆數": len(nav),
        }
        for period_label, period_years in PERIODS.items():
            ret = calc_period_return(nav, period_years)
            row[period_label] = ret
            # debug: 記錄每個期間的起訖淨值
            end_dt = nav.index[-1]
            start_dt = end_dt - timedelta(days=int(period_years * 365))
            subset = nav[nav.index >= start_dt]
            if len(subset) >= 5:
                fund_debug[f"{period_label}_起始"] = subset.index[0].strftime("%Y-%m-%d")
                fund_debug[f"{period_label}_起始淨值"] = round(float(subset.iloc[0]), 4)
                fund_debug[f"{period_label}_累積%"] = ret
        rows.append(row)
        debug_info.append(fund_debug)
        progress.progress((i+1)/len(selected_tickers), text=f"已讀取：{fund_name}")
    progress.empty()

    # 顯示 debug 資訊（可展開）
    if debug_info:
        with st.expander("🔍 資料驗證（點擊展開確認起訖日期與淨值）"):
            st.dataframe(pd.DataFrame(debug_info), use_container_width=True, hide_index=True)

    return pd.DataFrame(rows) if rows else None

st.set_page_config(page_title="匯率避險分析工具", page_icon="💱", layout="wide")

st.title("💱 基金投組匯率避險分析工具")
st.caption("計算基金報酬率能抵銷多少台幣升值幅度｜台北富邦銀行財富管理")

PERIODS = {
    "半年": 0.5,
    "1年": 1,
    "2年": 2,
    "3年": 3,
    "5年": 5,
    "7年": 7,
    "10年": 10,
}

# ── Sidebar 設定 ──────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 投組設定")

    usd_amount = st.number_input("美元資產（台幣萬元）", min_value=100, max_value=10000, value=950, step=50)
    exchange_rate = st.number_input("換匯匯率（台幣/美元）", min_value=25.0, max_value=40.0, value=32.0, step=0.1, format="%.2f")
    fund_amount = st.number_input("基金投資金額（台幣萬元）", min_value=10, max_value=1000, value=50, step=10)

    usd_return = st.number_input("美元資產年化報酬率（%）", min_value=0.0, value=0.0, step=0.1, format="%.1f")
    if usd_return > 0:
        st.caption(f"✅ 美元資產有生息，損益平衡匯率會更低（保護力更強）")

    usd_in_usd = usd_amount / exchange_rate
    st.info(f"💵 美元資產：{usd_in_usd:.2f} 萬美元")
    st.info(f"📊 總資產：{usd_amount + fund_amount:.0f} 萬台幣")

    st.divider()
    st.subheader("📋 基金清單輸入")
    input_mode = st.radio("輸入方式", ["Google Drive 自動抓", "手動輸入", "上傳CSV/Excel"])

# ── 基金資料輸入 ───────────────────────────────────────────
funds_df = None

if input_mode == "Google Drive 自動抓":
    st.subheader("☁️ 從 Google Drive 自動讀取基金淨值")
    st.caption("系統會自動從 Drive 抓取淨值並計算各期間年化報酬率（累加法）")

    selected_tickers = st.multiselect(
        "選擇基金（可多選）",
        options=list(FUND_DB.keys()),
        default=["F0HKG05X22_FO", "F00001EBH4_FO"],
        format_func=lambda x: FUND_DB[x]
    )

    if selected_tickers:
        if st.button("🔄 從 Google Drive 讀取資料", type="primary"):
            try:
                funds_df = fetch_fund_returns_from_drive(selected_tickers)
                if funds_df is not None:
                    st.session_state["drive_funds_df"] = funds_df
                    st.success(f"✅ 成功讀取 {len(funds_df)} 檔基金")
            except Exception as e:
                st.error(f"讀取失敗：{e}｜請確認 GOOGLE_CREDENTIALS 已設定在 Streamlit Secrets")

    if "drive_funds_df" in st.session_state:
        funds_df = st.session_state["drive_funds_df"]
        st.dataframe(funds_df, use_container_width=True)

elif input_mode == "手動輸入":
    st.subheader("✏️ 手動輸入基金資料")
    st.caption("請輸入各期間的年化報酬率（%）｜請輸入各期間的累積總報酬率（%），留空代表無資料")

    default_funds = [
        {"基金名稱": "安聯台灣科技 (ACDD04)", "半年": None, "1年": None, "2年": None, "3年": None, "5年": None, "7年": None, "10年": None},
        {"基金名稱": "元大全球優質龍頭平衡 (ACYT168)", "半年": None, "1年": None, "2年": None, "3年": None, "5年": None, "7年": None, "10年": None},
    ]

    edited_df = st.data_editor(
        pd.DataFrame(default_funds),
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "基金名稱": st.column_config.TextColumn("基金名稱", width="large"),
            **{p: st.column_config.NumberColumn(p, min_value=-100.0, format="%.1f%%") for p in PERIODS.keys()}
        }
    )
    funds_df = edited_df

else:
    st.subheader("📁 上傳基金資料")
    st.caption("檔案需包含欄位：基金名稱、半年、1年、2年、3年、5年、7年、10年（年化報酬率%）")

    template_df = pd.DataFrame([
        {"基金名稱": "範例基金A", "半年": 4.2, "1年": 6.8, "2年": 5.1, "3年": 4.5, "5年": 5.3, "7年": 5.8, "10年": 6.2},
        {"基金名稱": "範例基金B", "半年": "", "1年": 8.2, "2年": 6.3, "3年": 5.9, "5年": 7.1, "7年": 7.5, "10年": 8.0},
    ])
    buf = BytesIO()
    template_df.to_excel(buf, index=False)
    st.download_button("⬇️ 下載Excel範本", buf.getvalue(), "基金範本.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    uploaded = st.file_uploader("上傳檔案", type=["xlsx", "csv"])
    if uploaded:
        if uploaded.name.endswith(".csv"):
            funds_df = pd.read_csv(uploaded)
        else:
            funds_df = pd.read_excel(uploaded)
        st.success(f"✅ 成功載入 {len(funds_df)} 檔基金")
        st.dataframe(funds_df, use_container_width=True)

# ── 計算邏輯 ───────────────────────────────────────────────
def calc_breakeven(fund_cumulative_pct, period_years, fund_amount_twd, usd_amount_twd, usd_in_usd_val, usd_return_pct=0.0):
    """
    fund_cumulative_pct : 該期間累積總報酬%（例如2年漲了60%就填60）
    period_years        : 年數（只用於計算美元資產累積生息）
    usd_return_pct      : 美元資產年化報酬率%（線性累加）

    邏輯：
    - 基金累積獲利 = 本金 × 累積總報酬%（直接用，不乘年數）
    - 美元累積生息 = 美元數量 × 年化報酬率 × 年數（線性）
    - 損益平衡匯率 = (美元台幣成本 - 基金累積獲利) / 美元期末數量
    """
    # 基金累積獲利（直接用累積%，不乘年數）
    fund_profit = fund_amount_twd * (fund_cumulative_pct / 100)
    fund_final  = fund_amount_twd + fund_profit

    # 美元資產期末美元數（年化報酬線性累加）
    usd_final_usd = usd_in_usd_val * (1 + (usd_return_pct / 100) * period_years)

    if usd_final_usd <= 0:
        return None

    remaining = usd_amount_twd - fund_profit
    breakeven_rate = remaining / usd_final_usd
    appreciation_pct = (exchange_rate - breakeven_rate) / exchange_rate * 100

    return {
        "損益平衡匯率": round(breakeven_rate, 2),
        "可承受升值幅度(%)": round(appreciation_pct, 2),
        "基金累積獲利(萬)": round(fund_profit, 1),
        "基金期末總值(萬)": round(fund_final, 1),
        "美元資產期末(萬美元)": round(usd_final_usd, 2),
    }

# ── 主要輸出 ───────────────────────────────────────────────
if funds_df is not None and len(funds_df) > 0:
    st.divider()
    st.subheader("📊 分析結果")

    # 整理結果
    results = []
    chart_data = {}  # fund_name -> list of (period, breakeven_rate)

    for _, row in funds_df.iterrows():
        fund_name = row.get("基金名稱", "未命名")
        fund_row = {"基金名稱": fund_name}
        chart_data[fund_name] = []

        for period_label, period_years in PERIODS.items():
            val = row.get(period_label)
            if pd.isna(val) or val == "" or val is None:
                fund_row[f"{period_label}_損益平衡匯率"] = None
                fund_row[f"{period_label}_可承受升值(%)"] = None
                chart_data[fund_name].append((period_label, None))
            else:
                res = calc_breakeven(float(val), period_years, fund_amount, usd_amount, usd_in_usd, usd_return)
                fund_row[f"{period_label}_損益平衡匯率"] = res["損益平衡匯率"]
                fund_row[f"{period_label}_可承受升值(%)"] = res["可承受升值幅度(%)"]
                chart_data[fund_name].append((period_label, res["損益平衡匯率"]))

        results.append(fund_row)

    results_df = pd.DataFrame(results)

    # ── Tab 顯示 ──────────────────────────────────────────
    tab1, tab2, tab3 = st.tabs(["📋 損益平衡匯率表", "📈 折線圖", "⬇️ 下載報告"])

    with tab1:
        # 損益平衡匯率表
        st.markdown("#### 損益平衡匯率（台幣/美元）")
        st.caption(f"換匯匯率：{exchange_rate} ｜ 美元資產：{usd_amount}萬（年化報酬 {usd_return}%）｜ 基金投資：{fund_amount}萬")

        display_rows = []
        for _, row in funds_df.iterrows():
            fund_name = row.get("基金名稱", "未命名")
            r1 = {"基金名稱": fund_name, "指標": "累積報酬率(%)"}
            r2 = {"基金名稱": "", "指標": "損益平衡匯率"}
            r3 = {"基金名稱": "", "指標": "可承受升值幅度(%)"}
            for period_label in PERIODS.keys():
                val = row.get(period_label)
                if pd.isna(val) or val == "" or val is None:
                    r1[period_label] = "-"
                    r2[period_label] = "-"
                    r3[period_label] = "-"
                else:
                    res = calc_breakeven(float(val), PERIODS[period_label], fund_amount, usd_amount, usd_in_usd, usd_return)
                    r1[period_label] = f"{float(val):.1f}%"
                    r2[period_label] = f"{res['損益平衡匯率']:.2f}"
                    r3[period_label] = f"{res['可承受升值幅度(%)']:.1f}%"
            display_rows.extend([r1, r2, r3, {"基金名稱": "─" * 20, "指標": "", **{p: "" for p in PERIODS.keys()}}])

        display_df = pd.DataFrame(display_rows)
        st.dataframe(display_df, use_container_width=True, hide_index=True)

        # 警戒線說明
        st.info(f"📌 換匯匯率 {exchange_rate}，若損益平衡匯率 < 28 代表台幣須升值超過警戒水位，避險效果有限")

    with tab2:
        st.markdown("#### 各基金損益平衡匯率走勢")

        fig = go.Figure()
        colors = ["#c9a84c", "#3b82f6", "#22c55e", "#ef4444", "#a855f7", "#f97316", "#06b6d4"]

        for i, (fund_name, data_points) in enumerate(chart_data.items()):
            # 實際線（含美元報酬率設定）
            xs = [p for p, v in data_points if v is not None]
            ys = [v for _, v in data_points if v is not None]
            if xs:
                fig.add_trace(go.Scatter(
                    x=xs, y=ys, mode="lines+markers",
                    name=f"{fund_name}（美元{usd_return}%）",
                    line=dict(color=colors[i % len(colors)], width=2.5),
                    marker=dict(size=8),
                    hovertemplate=f"<b>{fund_name}</b>（美元{usd_return}%）<br>%{{x}}: %{{y:.2f}} 元<extra></extra>"
                ))

            # 對比線：美元資產0%（虛線）
            if usd_return > 0:
                row_data = funds_df[funds_df["基金名稱"] == fund_name]
                if not row_data.empty:
                    xs0, ys0 = [], []
                    for period_label, period_years in PERIODS.items():
                        val = row_data.iloc[0].get(period_label)
                        if not (pd.isna(val) or val == "" or val is None):
                            res0 = calc_breakeven(float(val), period_years, fund_amount, usd_amount, usd_in_usd, 0.0)
                            xs0.append(period_label)
                            ys0.append(res0["損益平衡匯率"])
                    if xs0:
                        fig.add_trace(go.Scatter(
                            x=xs0, y=ys0, mode="lines+markers",
                            name=f"{fund_name}（美元0%）",
                            line=dict(color=colors[i % len(colors)], width=1.5, dash="dot"),
                            marker=dict(size=6, symbol="circle-open"),
                            opacity=0.6,
                            hovertemplate=f"<b>{fund_name}</b>（美元0%）<br>%{{x}}: %{{y:.2f}} 元<extra></extra>"
                        ))

        # 參考線
        fig.add_hline(y=exchange_rate, line_dash="dot", line_color="#22c55e", annotation_text=f"起始匯率 {exchange_rate}", annotation_position="top left")
        fig.add_hline(y=28, line_dash="dash", line_color="#ef4444", annotation_text="28元 警戒線", annotation_position="bottom right")
        fig.add_hline(y=25, line_dash="dash", line_color="#f97316", annotation_text="25元 極端情境", annotation_position="bottom right")

        fig.update_layout(
            plot_bgcolor="#0f1729",
            paper_bgcolor="#0f1729",
            font=dict(color="#e2e8f0"),
            legend=dict(bgcolor="#1e2d45", bordercolor="#334155"),
            xaxis=dict(title="投資期間", gridcolor="#1e2d45"),
            yaxis=dict(title="損益平衡匯率（台幣/美元）", gridcolor="#1e2d45", range=[23, exchange_rate + 1]),
            height=500,
            hovermode="x unified"
        )
        st.plotly_chart(fig, use_container_width=True)

        if usd_return > 0:
            st.caption(f"實線 = 美元資產有 {usd_return}% 報酬；虛線 = 美元資產0%（對比）。實線越低於虛線，代表美元生息效果越顯著。")
        else:
            st.caption("損益平衡匯率越低，代表基金獲利越多、能抵銷更大幅度的台幣升值。低於28元進入警戒區。設定美元資產報酬率可顯示對比虛線。")

    with tab3:
        st.markdown("#### 下載Excel分析報告")

        def generate_excel():
            wb = openpyxl.Workbook()

            # ── Sheet 1: 摘要 ──
            ws1 = wb.active
            ws1.title = "損益平衡分析"

            header_fill = PatternFill("solid", start_color="1E3A5F")
            gold_fill = PatternFill("solid", start_color="C9A84C")
            green_fill = PatternFill("solid", start_color="1A4731")
            red_fill = PatternFill("solid", start_color="4B1A1A")
            alt_fill = PatternFill("solid", start_color="111827")
            white_font = Font(color="FFFFFF", bold=True, name="Arial")
            gold_font = Font(color="C9A84C", bold=True, name="Arial")
            normal_font = Font(color="E2E8F0", name="Arial")
            thin = Side(style="thin", color="1E2D45")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center")

            # 標題
            ws1.merge_cells("A1:J1")
            ws1["A1"] = "基金投組匯率避險分析報告"
            ws1["A1"].font = Font(color="C9A84C", bold=True, size=14, name="Arial")
            ws1["A1"].fill = PatternFill("solid", start_color="0A0F1E")
            ws1["A1"].alignment = center

            # 參數列
            ws1.merge_cells("A2:J2")
            ws1["A2"] = f"換匯匯率：{exchange_rate} ｜ 美元資產：{usd_amount}萬台幣（{usd_in_usd:.2f}萬美元）｜ 基金投資：{fund_amount}萬台幣"
            ws1["A2"].font = Font(color="94A3B8", size=10, name="Arial")
            ws1["A2"].fill = PatternFill("solid", start_color="0A0F1E")
            ws1["A2"].alignment = center

            period_labels = list(PERIODS.keys())

            # 表頭
            headers = ["基金名稱", "指標"] + period_labels
            for col_idx, h in enumerate(headers, 1):
                cell = ws1.cell(row=4, column=col_idx, value=h)
                cell.font = white_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border

            row = 5
            for fund_idx, fund_row_data in funds_df.iterrows():
                fund_name = fund_row_data.get("基金名稱", "未命名")
                bg = PatternFill("solid", start_color="111827") if fund_idx % 2 == 0 else PatternFill("solid", start_color="0D1520")

                for metric_label, metric_key in [("年化報酬率(%)", "return"), ("損益平衡匯率", "breakeven"), ("可承受升值幅度(%)", "appreciation")]:
                    ws1.cell(row=row, column=1, value=fund_name if metric_label == "年化報酬率(%)" else "").font = Font(color="C9A84C", bold=True, name="Arial")
                    ws1.cell(row=row, column=1).fill = bg
                    ws1.cell(row=row, column=1).border = border
                    ws1.cell(row=row, column=2, value=metric_label).font = Font(color="94A3B8", name="Arial", size=9)
                    ws1.cell(row=row, column=2).fill = bg
                    ws1.cell(row=row, column=2).border = border
                    ws1.cell(row=row, column=2).alignment = center

                    for col_idx, period_label in enumerate(period_labels, 3):
                        val = fund_row_data.get(period_label)
                        cell = ws1.cell(row=row, column=col_idx)
                        cell.fill = bg
                        cell.border = border
                        cell.alignment = center

                        if pd.isna(val) or val == "" or val is None:
                            cell.value = "-"
                            cell.font = Font(color="4B5563", name="Arial")
                        else:
                            res = calc_breakeven(float(val), PERIODS[period_label], fund_amount, usd_amount, usd_in_usd, usd_return)
                            if metric_key == "return":
                                cell.value = float(val) / 100
                                cell.number_format = "0.0%"
                                cell.font = Font(color="60A5FA", bold=True, name="Arial")
                            elif metric_key == "breakeven":
                                br = res["損益平衡匯率"]
                                cell.value = br
                                cell.number_format = "0.00"
                                if br >= 30:
                                    cell.font = Font(color="4ADE80", bold=True, name="Arial")
                                elif br >= 28:
                                    cell.font = Font(color="FACC15", bold=True, name="Arial")
                                else:
                                    cell.font = Font(color="F87171", bold=True, name="Arial")
                            elif metric_key == "appreciation":
                                cell.value = res["可承受升值幅度(%)"] / 100
                                cell.number_format = "0.0%"
                                cell.font = Font(color="A78BFA", name="Arial")
                    row += 1

                # 分隔行
                for col_idx in range(1, len(headers) + 1):
                    sep = ws1.cell(row=row, column=col_idx, value="")
                    sep.fill = PatternFill("solid", start_color="0A0F1E")
                row += 1

            # 欄寬
            ws1.column_dimensions["A"].width = 28
            ws1.column_dimensions["B"].width = 18
            for col_idx in range(3, 3 + len(period_labels)):
                ws1.column_dimensions[get_column_letter(col_idx)].width = 12
            ws1.row_dimensions[1].height = 28
            ws1.freeze_panes = "C5"

            # ── Sheet 2: 原始資料 ──
            ws2 = wb.create_sheet("基金原始資料")
            for col_idx, col_name in enumerate(funds_df.columns, 1):
                cell = ws2.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF", name="Arial")
                cell.fill = PatternFill("solid", start_color="1E3A5F")
                cell.alignment = center
                cell.border = border
            for r_idx, row_data in funds_df.iterrows():
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws2.cell(row=r_idx + 2, column=col_idx, value=val)
                    cell.font = Font(name="Arial")
                    cell.border = border
                    cell.alignment = center

            buf = BytesIO()
            wb.save(buf)
            return buf.getvalue()

        excel_bytes = generate_excel()
        st.download_button(
            label="⬇️ 下載Excel報告",
            data=excel_bytes,
            file_name="匯率避險分析報告.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        st.markdown("""
**Excel報告內容：**
- 📋 損益平衡分析表（含顏色標示：綠=安全、黃=注意、紅=警戒）
- 📁 基金原始資料頁
        """)

else:
    st.info("👈 請先在左側設定投組參數，並輸入基金資料")
    st.markdown("""
    **使用說明：**
    1. 在左側設定美元資產金額、換匯匯率、基金投資金額
    2. 選擇手動輸入或上傳CSV/Excel
    3. 輸入各基金在不同期間的**年化報酬率（%）**
    4. 系統自動計算各期間的**損益平衡匯率**
    
    **損益平衡匯率** = 基金獲利剛好補足美元匯損時的台幣匯率
    - 若台幣升值到此匯率以下 → 整體投組開始虧損
    - 數字越低 → 基金保護力越強
    """)
